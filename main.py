from datetime import date
from datetime import datetime
from io import BytesIO
import os
from flask import Flask, Response, request, redirect, render_template, url_for, flash, jsonify, send_from_directory, send_file
import sqlite3
import openpyxl

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

import webbrowser
import threading

app = Flask(__name__)
app.secret_key = os.urandom(24)

DATABASE = './database.db'
UPLOAD_FOLDER = './uploads/'


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def create_database():

    db = get_db()
    c = db.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS students (
              student_id INTEGER PRIMARY KEY UNIQUE, 
              title TEXT NOT NULL,
              f_name TEXT NOT NULL,
              l_name TEXT NOT NULL,
              f_eng_name TEXT,
              l_eng_name TEXT,
              n_name TEXT,
              address TEXT,
              dob TEXT,
              age INTEGER,
              job TEXT,
              email TEXT,
              phone_num TEXT,
              instruments TEXT,

              family_title TEXT,
              family_f_name TEXT,
              family_l_name TEXT,
              family_relationship TEXT,

              register_date TEXT,
              
              created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS teachers (
              teacher_id INTEGER PRIMARY KEY UNIQUE, 
              title TEXT NOT NULL,
              f_name TEXT NOT NULL,
              l_name TEXT NOT NULL,
              n_name TEXT,
              address TEXT,
              dob TEXT,
              job TEXT,
              email TEXT,
              phone_num TEXT,

              payment_ratio REAL,

              created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS subjects (
              subject_id INTEGER PRIMARY KEY,
              subject_name TEXT
    )''')

    c.execute(
        '''INSERT INTO subjects (subject_name) VALUES ("เปียโนเด็กพื้นฐาน"), ("เปียโนเด็กเดี่ยว (30 นาที)"), ("เปียโนเด็กเดี่ยว (45 นาที)"), ("เปียโนเด็กเดี่ยว (1 ชั่วโมง)"), ("เปียโน"), ("ไวโอลิน"), ("กีต้าร์"), ("ดนตรีไทย")''')

    c.execute('''CREATE TABLE IF NOT EXISTS levels (
              level_id INTEGER PRIMARY KEY,
              level_name TEXT
    )''')

    c.execute(
        '''INSERT INTO levels (level_name) VALUES ("Beginner"), ("Intermediate"), ("Advanced")''')

    c.execute('''CREATE TABLE IF NOT EXISTS registrations (
              registration_id INTEGER PRIMARY KEY,
              student_id INTEGER,
              subject_id INTEGER,
              level_id INTEGER,
              teacher_id INTEGER,
              time_left INTEGER,
              times INTEGER,
              schedule TEXT NOT NULL,
              created_at DATETIME DEFAULT CURRENT_TIMESTAMP,

              FOREIGN KEY (student_id) REFERENCES students(student_id),
              FOREIGN KEY (subject_id) REFERENCES subjects(subject_id),
              FOREIGN KEY (level_id) REFERENCES levels(level_id),
              FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id)
              
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS attendances (
              attendance_id INTEGER PRIMARY KEY,
              student_id INTEGER, 
              subject_id INTEGER,
              level_id INTEGER,
              date DATE,
              created_at DATETIME DEFAULT CURRENT_TIMESTAMP,

              FOREIGN KEY (student_id) REFERENCES students(student_id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS prices (
              price_id INTEGER PRIMARY KEY,
              subject_id INTEGER,
              level_id INTEGER,
              price INTEGER,

              FOREIGN KEY (subject_id) REFERENCES subjects(subject_id),
              FOREIGN KEY (level_id) REFERENCES levels(level_id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS payments (
        payment_id INTEGER PRIMARY KEY,
        teacher_id INTEGER,
        attendance_id INTEGER,
        student_id INTEGER,
        subject_id INTEGER,
        level_id INTEGER,
        amount INTEGER,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
              
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id),
        FOREIGN KEY (attendance_id) REFERENCES attendances(attendance_id),
        FOREIGN KEY (student_id) REFERENCES students(student_id),
        FOREIGN KEY (subject_id) REFERENCES subjects(subject_id),
        FOREIGN KEY (level_id) REFERENCES levels(level_id)
    )''')

    # Data for testing
    c.execute('''INSERT INTO teachers (teacher_id, title, f_name, l_name, payment_ratio) VALUES (987, "นาย", "Teerapath", "Sattabongkot", 0.5)''')

    # เปียโนเด็กพื้นฐาน
    c.execute('''INSERT INTO prices (subject_id, level_id, price) VALUES (1, 1, 2800), (1, 2, 3200), (1, 3, 3400)''')

    # Violin
    c.execute('''INSERT INTO prices (subject_id, level_id, price) VALUES (2, 1, 2800), (2, 2, 3200), (2, 3, 3400)''')

    # Guitar
    c.execute('''INSERT INTO prices (subject_id, level_id, price) VALUES (3, 1, 2800), (3, 2, 3200), (3, 3, 3400)''')

    # ดนตรีไทย
    c.execute('''INSERT INTO prices (subject_id, level_id, price) VALUES (4, 1, 2800), (4, 2, 3200), (4, 3, 3400)''')

    db.commit()
    db.close()


@app.route('/favicon.ico', methods=['GET', 'POST'])
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'favicon.ico', mimetype='image/vnd.microsoft.icon')


@app.route("/", methods=['GET', 'POST'])
def main():
    if not os.path.exists(DATABASE):
        os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
        create_database()

    # Ensure upload folder exists
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)

    if request.method == 'POST':
        username_login = request.form['username_login']

        if username_login == "admin":
            return redirect(url_for('students'))
        else:
            flash("ชื่อผู้ใช้งานไม่ถูกต้อง", "danger")
            return render_template('index.html')

    return render_template('index.html')


@app.route("/students", methods=['GET', 'POST'])
@app.route("/students/<student_id>", methods=['GET', 'POST'])
def students(student_id=None):

    if student_id:

        # Get student information
        db = get_db()
        c = db.cursor()
        c.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
        student = c.fetchone()
        db.close()

        return render_template('student_information.html', student=student)

    today = date.today().strftime('%Y-%m-%d')

    db = get_db()
    c = db.cursor()
    c.execute("""
        SELECT 
            students.student_id, 
            students.title, 
            students.f_name, 
            students.l_name, 
            students.n_name, 
            students.email, 
            students.phone_num,
            CASE 
                WHEN attendances.date IS NOT NULL THEN 1
                ELSE 0
            END AS is_attended
        FROM students
        LEFT JOIN attendances 
            ON students.student_id = attendances.student_id 
            AND attendances.date = ?
    """, (today,))
    students = c.fetchall()
    db.close()

    return render_template('students.html', students=students)


@app.route("/teachers", methods=['GET', 'POST'])
@app.route("/teachers/<teacher_id>", methods=['GET', 'POST'])
def teachers(teacher_id=None):

    if teacher_id:

        # Get teacher information
        db = get_db()
        c = db.cursor()
        c.execute("SELECT * FROM teachers WHERE teacher_id = ?", (teacher_id,))
        teacher = c.fetchone()
        db.close()

        return render_template('teacher_information.html', teacher=teacher)

    db = get_db()
    c = db.cursor()
    c.execute(
        "SELECT teacher_id, f_name, l_name, n_name, email, phone_num FROM teachers")
    teachers = c.fetchall()
    db.close()

    # Convert to dict
    teachers = [dict(teacher) for teacher in teachers]
    for teacher in teachers:

        teacher_id = teacher['teacher_id']

        # Get summation of payment of the teachers of each month
        db = get_db()
        c = db.cursor()
        c.execute("""
            SELECT
                SUM(payments.amount) AS total_amount
            FROM payments 
            JOIN attendances ON payments.attendance_id = attendances.attendance_id
            WHERE payments.teacher_id = ? and strftime('%Y-%m', attendances.date) = strftime('%Y-%m', 'now')
        """, (teacher_id,))
        payment = c.fetchone()
        db.close()

        payment_amount = payment['total_amount'] if payment['total_amount'] is not None else 0
        teacher['payment'] = payment_amount

    # Group by month, get summation of payment
    db = get_db()
    c = db.cursor()
    c.execute(
        '''SELECT strftime('%Y-%m', attendances.date) AS month_year, SUM(payments.amount) AS total_amount 
        FROM payments 
        JOIN attendances ON payments.attendance_id = attendances.attendance_id
        GROUP BY month_year''')
    total_monthly_payments = c.fetchall()
    db.close()

    return render_template('teachers.html', teachers=teachers, total_monthly_payments=total_monthly_payments)


@app.route("/subjects", methods=['GET', 'POST'])
def subjects():
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM subjects")
    subjects = c.fetchall()
    db.close()

    return render_template('subjects.html', subjects=subjects)


@app.route("/levels", methods=['GET', 'POST'])
def levels():

    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM levels")
    levels = c.fetchall()
    db.close()

    return render_template('levels.html', levels=levels)


@app.route("/add/<category>", methods=['GET', 'POST'])
def add(category):

    if request.method == 'POST':

        if category == "student":

            register_date = request.form['register_date']
            if not register_date:
                flash("กรุณากรอกวันที่ลงทะเบียน", "danger")
                return redirect(url_for('students'))

            student_id = request.form['student_id']
            title = request.form['title']
            f_name = request.form['f_name']
            l_name = request.form['l_name']

            f_eng_name = request.form['f_eng_name']
            l_eng_name = request.form['l_eng_name']

            n_name = request.form['n_name']
            address = request.form['address']

            dob = request.form['dob']
            if not register_date:
                flash("กรุณากรอกวันเกิด", "danger")
                return redirect(url_for('students'))

            age = request.form['age']

            job = request.form['job']
            email = request.form['email']
            phone_num = request.form['phone_num']
            instruments = request.form['instruments']

            family_title = request.form['family_title']
            family_f_name = request.form['family_f_name']
            family_l_name = request.form['family_l_name']
            family_relationship = request.form['family_relationship']

            db = get_db()
            c = db.cursor()
            c.execute('''INSERT INTO students (student_id, title, f_name, l_name, f_eng_name, l_eng_name, n_name, address, dob, age, job, email, phone_num, instruments, family_title, family_f_name, family_l_name, family_relationship, register_date) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (student_id, title, f_name, l_name, f_eng_name, l_eng_name, n_name, address, dob, age, job, email, phone_num, instruments, family_title, family_f_name, family_l_name, family_relationship, register_date))
            db.commit()
            db.close()

            flash("เพิ่มนักเรียนสําเร็จ", "success")
            return redirect(url_for('students'))

        elif category == "teacher":
            teacher_id = request.form['teacher_id']
            title = request.form['title']
            f_name = request.form['f_name']
            l_name = request.form['l_name']
            n_name = request.form['n_name']
            address = request.form['address']
            dob = request.form['dob']
            job = request.form['job']
            email = request.form['email']
            phone_num = request.form['phone_num']

            payment_ratio = request.form['payment_ratio']

            db = get_db()
            c = db.cursor()
            c.execute('''INSERT INTO teachers (teacher_id, title, f_name, l_name, n_name, address, dob, job, email, phone_num, payment_ratio) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (teacher_id, title, f_name, l_name, n_name, address, dob, job, email, phone_num, payment_ratio))
            db.commit()
            db.close()

            flash("เพิ่มอาจารย์สําเร็จ", "success")
            return redirect(url_for('teachers'))

        elif category == "subject":

            subject_name = request.form['subject_name']

            db = get_db()
            c = db.cursor()
            c.execute('''INSERT INTO subjects (subject_name) 
                    VALUES (?)''', (subject_name,))
            db.commit()
            db.close()

            flash("เพิ่มวิชาสําเร็จ", "success")
            return redirect(url_for('subjects'))

        elif category == "level":

            level_name = request.form['level_name']

            db = get_db()
            c = db.cursor()
            c.execute('''INSERT INTO levels (level_name)
                    VALUES (?)''', (level_name,))
            db.commit()
            db.close()

            flash("เพิ่มระดับชั้นสําเร็จ", "success")
            return redirect(url_for('levels'))

        elif category == "price":

            subject_id = request.form['subject_id']
            level_id = request.form['level_id']
            price = request.form['price']

            db = get_db()
            c = db.cursor()
            c.execute("INSERT INTO prices (subject_id, level_id, price) VALUES (?, ?, ?)",
                      (subject_id, level_id, price))
            db.commit()
            db.close()

            flash("เพิ่มราคาสําเร็จ", "success")
            return redirect(url_for('prices', subject_id=subject_id))

        elif category == "attendance":

            student_id = request.form['student_id']

            # Get subject_id and level_id from registrations according to student_id
            db = get_db()
            c = db.cursor()
            c.execute(
                "SELECT subject_id, level_id FROM registrations WHERE student_id = ?", (student_id,))
            rows = c.fetchall()
            db.close()

            for row in rows:
                add_attendance(student_id, row[0], row[1], date.today())

            flash("เช็คชื่อเรียบร้อยแล้ว", "success")
            return redirect(url_for('students'))

        elif category == "registration":

            student_id = request.args.get('student_id')

            subject_id = request.form['subject_id']
            level_id = request.form['level_id']
            teacher_id = request.form['teacher_id']
            times = request.form['times']
            schedule = request.form['schedule']

            if int(times) % 4 != 0:
                flash("จํานวนครั้งที่สอนต้องหาร 4 ลงตัว", "warning")
                return redirect(url_for('registrations', student_id=student_id))

            db = get_db()
            c = db.cursor()
            c.execute("INSERT INTO registrations (student_id, subject_id, level_id, teacher_id, time_left, times, schedule) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      (student_id, subject_id, level_id, teacher_id, times, times, schedule))
            db.commit()
            db.close()

            flash("เพิ่มวิชาที่เรียนเรียบร้อยแล้ว", "success")

            return redirect(url_for('receipt', student_id=student_id, subject_id=subject_id, level_id=level_id))

    else:
        return redirect(url_for('students'))


@app.route("/update/<category>", methods=['GET', 'POST'])
def update(category):

    if request.method == 'POST':
        if category == "student":

            student_id = request.args.get('student_id')

            title = request.form['title']
            f_name = request.form['f_name']
            l_name = request.form['l_name']
            n_name = request.form['n_name']
            address = request.form['address']
            dob = request.form['dob']
            job = request.form['job']
            email = request.form['email']
            phone_num = request.form['phone_num']
            instruments = request.form['instruments']

            family_title = request.form['family_title']
            family_f_name = request.form['family_f_name']
            family_l_name = request.form['family_l_name']
            family_relationship = request.form['family_relationship']

            db = get_db()
            c = db.cursor()
            c.execute("""UPDATE students SET title=?, f_name=?, l_name=?, n_name=?, address=?, dob=?, job=?, email=?, phone_num=?, instruments=?, family_title=?, family_f_name=?, family_l_name=?, family_relationship=? 
                      WHERE student_id=?""", (title, f_name, l_name, n_name, address, dob, job, email, phone_num, instruments, family_title, family_f_name, family_l_name, family_relationship, student_id))
            db.commit()
            db.close()

            return redirect(url_for('students'))

        elif category == "teacher":

            teacher_id = request.args.get('teacher_id')

            title = request.form['title']
            f_name = request.form['f_name']
            l_name = request.form['l_name']
            n_name = request.form['n_name']
            address = request.form['address']
            dob = request.form['dob']
            job = request.form['job']
            email = request.form['email']
            phone_num = request.form['phone_num']

            payment_ratio = request.form['payment_ratio']

            db = get_db()
            c = db.cursor()
            c.execute("""UPDATE teachers SET title=?, f_name=?, l_name=?, n_name=?, address=?, dob=?, job=?, email=?, phone_num=?, payment_ratio=? 
                      WHERE teacher_id=?""", (title, f_name, l_name, n_name, address, dob, job, email, phone_num, payment_ratio, teacher_id))
            db.commit()
            db.close()

            return redirect(url_for('teachers'))

    else:
        return redirect(url_for('students'))


@app.route("/delete/<category>", methods=['GET', 'POST'])
def delete(category=None):

    if category:

        if category == "student":

            student_id = request.args.get('student_id')

            db = get_db()
            c = db.cursor()
            c.execute("DELETE FROM students WHERE student_id = ?", (student_id,))
            db.commit()
            db.close()

            flash("ลบนักเรียนเรียบร้อยแล้ว", "success")

            return redirect(url_for('students'))

        elif category == "teacher":

            teacher_id = request.args.get('teacher_id')

            db = get_db()
            c = db.cursor()
            c.execute("DELETE FROM teachers WHERE teacher_id = ?", (teacher_id,))
            db.commit()
            db.close()

            flash("ลบอาจารย์เรียบร้อยแล้ว", "success")

            return redirect(url_for('teachers'))

        elif category == "subject":

            subject_id = request.args.get('subject_id')

            db = get_db()
            c = db.cursor()
            c.execute("DELETE FROM subjects WHERE subject_id = ?", (subject_id,))
            db.commit()
            db.close()

            flash("ลบรายวิชาเรียบร้อยแล้ว", "success")

            return redirect(url_for('subjects'))

        elif category == "level":

            level_id = request.args.get('level_id')

            db = get_db()
            c = db.cursor()
            c.execute("DELETE FROM levels WHERE level_id = ?", (level_id,))
            db.commit()
            db.close()

            flash("ลบระดับชั้นเรียบร้อยแล้ว", "success")

            return redirect(url_for('levels'))

        elif category == "registration":

            registration_id = request.args.get('registration_id')

            db = get_db()
            c = db.cursor()
            c.execute(
                "DELETE FROM registrations WHERE registration_id = ?", (registration_id,))
            db.commit()
            db.close()

            flash("ลบการลงวิชาเรียบร้อยแล้ว", "success")

            return redirect(url_for('students'))

        elif category == "attendance":

            student_id = request.args.get('student_id')
            subject_id = request.args.get('subject_id')
            level_id = request.args.get('level_id')
            attendance_id = request.args.get('attendance_id')

            delete_attendance(student_id, subject_id, level_id, attendance_id)

            flash("ลบประวัติการเข้าเรียนเรียบร้อยแล้ว", "success")
            return redirect(url_for('students'))

        elif category == "price":

            price_id = request.args.get('price_id')
            subject_id = request.args.get('subject_id')

            db = get_db()
            c = db.cursor()
            c.execute("DELETE FROM prices WHERE price_id = ?", (price_id,))
            db.commit()
            db.close()

            flash("ลบเรียบร้อยแล้ว", "success")

            return redirect(url_for('prices', subject_id=subject_id))
    else:
        return redirect(url_for('students'))


@app.route("/registrations/<student_id>", methods=['GET', 'POST'])
def registrations(student_id):

    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
    student = c.fetchone()
    db.close()

    if not student:
        flash("ไม่มีนักเรียนไอดีนี้", "warning")
        return redirect(url_for('students'))

    # Get all subjects and its levels
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM subjects")
    subjects = c.fetchall()
    db.close()

    if len(subjects) == 0:
        flash("ไม่มีรายวิชาที่สอน", "warning")
        return redirect(url_for('students'))

    # Get all teachers
    db = get_db()
    c = db.cursor()
    c.execute("SELECT teacher_id, title, f_name, l_name, n_name FROM teachers")
    teachers = c.fetchall()
    db.close()

    # Get registrations according to student_id
    db = get_db()
    c = db.cursor()
    c.execute('''SELECT registrations.registration_id, students.student_id, subjects.subject_id, levels.level_id, subjects.subject_name, levels.level_name, teachers.n_name, registrations.time_left, registrations.times, registrations.schedule 
              FROM registrations 
              JOIN subjects ON registrations.subject_id = subjects.subject_id
              JOIN levels ON registrations.level_id = levels.level_id
              JOIN teachers ON registrations.teacher_id = teachers.teacher_id
              JOIN students ON registrations.student_id = students.student_id
              WHERE registrations.student_id = ? 
              GROUP BY registrations.registration_id''', (student_id,))
    rows = c.fetchall()
    db.close()

    return render_template('registrations.html', student=student, rows=rows, subjects=subjects, teachers=teachers)


@app.route('/get_prices/<int:subject_id>', methods=['GET', 'POST'])
def get_levels(subject_id):
    db = get_db()
    c = db.cursor()
    # Get levels based on the selected subject
    c.execute('''SELECT levels.level_id, levels.level_name 
              FROM prices 
              JOIN levels ON prices.level_id = levels.level_id 
              WHERE subject_id = ?''', (subject_id,))
    rows = c.fetchall()
    db.close()

    # Convert rows to a list of dictionaries
    levels = [{"level_id": row[0], "level_name": row[1]} for row in rows]

    return jsonify(levels)


def add_attendance(student_id, subject_id, level_id, attend_date: date):

    # Check if attend_date already exist in attendances
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM attendances WHERE student_id = ? AND subject_id = ? AND level_id = ? AND date = ?",
              (student_id, subject_id, level_id, attend_date))
    row = c.fetchone()
    db.close()

    if row:
        flash("วันที่ " + str(attend_date) + " เช็คชื่อไปแล้ว", "warning")
        return

    # Add attendance
    db = get_db()
    c = db.cursor()
    c.execute("INSERT INTO attendances (student_id, subject_id, level_id, date) VALUES (?, ?, ?, ?)",
              (student_id, subject_id, level_id, attend_date))
    last_attendance_id = c.lastrowid
    db.commit()
    db.close()

    # Update student subject time left
    db = get_db()
    c = db.cursor()
    c.execute(
        '''UPDATE registrations SET time_left = time_left - 1 WHERE student_id = ? and subject_id = ? and level_id = ?''', (student_id, subject_id, level_id))
    db.commit()
    db.close()

    # Get teacher_id, subject_price, teacher_payment_ratio
    db = get_db()
    c = db.cursor()
    c.execute(
        """SELECT teachers.teacher_id, prices.price, teachers.payment_ratio FROM registrations
        JOIN teachers ON registrations.teacher_id = teachers.teacher_id
        JOIN prices ON registrations.subject_id = prices.subject_id AND registrations.level_id = prices.level_id
        WHERE registrations.student_id = ? AND registrations.subject_id = ? AND registrations.level_id = ?""", (student_id, subject_id, level_id))
    row = c.fetchone()
    db.commit()
    db.close()

    # Calculate money that teacher will get
    teacher_id = row[0]
    subject_price = row[1]
    teacher_payment_ratio = row[2]

    money = (subject_price * teacher_payment_ratio) // 4

    # Add transaction history to payments
    db = get_db()
    c = db.cursor()
    c.execute("INSERT INTO payments (teacher_id, attendance_id, student_id, subject_id, level_id, amount) VALUES (?, ?, ?, ?, ?, ?)",
              (teacher_id, last_attendance_id, student_id, subject_id, level_id, money))
    db.commit()
    db.close()


def delete_attendance(student_id, subject_id, level_id, attendance_id):

    # Remove transaction history
    db = get_db()
    c = db.cursor()
    c.execute("DELETE FROM payments WHERE attendance_id = ?",
              (attendance_id,))
    db.commit()
    db.close()

    # Delete attendance
    db = get_db()
    c = db.cursor()
    c.execute("DELETE FROM attendances WHERE attendance_id = ?", (attendance_id,))

    # Update student subject time left
    c.execute(
        '''UPDATE registrations SET time_left = time_left + 1 WHERE student_id = ? and subject_id = ? and level_id = ?''', (student_id, subject_id, level_id))
    db.commit()
    db.close()


@app.route("/attendances/<student_id>", methods=['GET', 'POST'])
def attendances(student_id):

    if request.method == 'POST':

        data = request.get_json()
        dates = data.get('dates', [])

        for attendance_date in dates:
            attendance_date = datetime.strptime(attendance_date, '%d/%m/%Y')

            # Convert year to Georgian year
            attendance_date = attendance_date.replace(
                year=attendance_date.year - 543).date()

            # Get subject_id and level_id from registrations according to student_id
            db = get_db()
            c = db.cursor()
            c.execute(
                "SELECT subject_id, level_id FROM registrations WHERE student_id = ?", (student_id,))
            rows = c.fetchall()
            db.close()

            for row in rows:
                add_attendance(student_id, row[0], row[1], attendance_date)

        return jsonify({'message': 'Success'})

    # Get student
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
    student = c.fetchone()
    db.close()

    if not student:
        flash("ไม่มีนักเรียนไอดีนี้", "warning")
        return redirect(url_for('students'))

    # Get attendance history
    db = get_db()
    c = db.cursor()
    c.execute('''SELECT attendance_id, date, student_id, subject_id, level_id 
              FROM attendances 
              WHERE student_id = ?''', (student_id,))
    attendances = c.fetchall()
    db.close()

    # Convert to dict
    attendances = [dict(attendance) for attendance in attendances]

    # Convert date to dd/mm/yyyy
    for attendance in attendances:
        date_obj = datetime.strptime(attendance['date'], '%Y-%m-%d')

        # Convert Georgian year to Thai year
        new_date_obj = date_obj.replace(year=date_obj.year + 543).date()

        attendance['date'] = new_date_obj.strftime('%d/%m/%Y')

    return render_template('attendances.html', student=student, attendances=attendances)


@app.route("/prices/<subject_id>", methods=['GET', 'POST'])
def prices(subject_id):

    db = get_db()
    c = db.cursor()
    c.execute('''SELECT * FROM prices 
              JOIN levels ON prices.level_id = levels.level_id 
              WHERE subject_id = ?''', (subject_id,))
    prices = c.fetchall()
    db.close()

    # Get subject id and name
    db = get_db()
    c = db.cursor()
    c.execute(
        "SELECT subject_id, subject_name FROM subjects WHERE subject_id = ?", (subject_id,))
    subject = c.fetchone()
    db.close()

    # Get levels
    db = get_db()
    c = db.cursor()
    c.execute("SELECT level_id, level_name FROM levels")
    levels = c.fetchall()
    db.close()

    return render_template('prices.html', prices=prices, subject=subject, levels=levels)


@app.route("/payments/<teacher_id>", methods=['GET', 'POST'])
def payments(teacher_id):

    # Get teacher
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM teachers WHERE teacher_id = ?", (teacher_id,))
    teacher = c.fetchone()
    db.close()

    # Get payments by teacher
    db = get_db()
    c = db.cursor()
    c.execute("""SELECT students.f_name, students.l_name, students.n_name, subjects.subject_name, levels.level_name, payments.amount, attendances.date FROM payments 
              JOIN students ON payments.student_id = students.student_id
              JOIN subjects ON payments.subject_id = subjects.subject_id
              JOIN levels ON payments.level_id = levels.level_id
              JOIN attendances ON payments.attendance_id = attendances.attendance_id
              WHERE teacher_id = ?""", (teacher_id,))
    payments = c.fetchall()
    db.close()

    # Combine payment in each month
    db = get_db()
    c = db.cursor()
    c.execute("""
        SELECT
            strftime('%Y-%m', attendances.date) AS month_year,
            SUM(payments.amount) AS total_amount
        FROM payments 
        JOIN attendances ON payments.attendance_id = attendances.attendance_id
        JOIN students ON payments.student_id = students.student_id
        JOIN subjects ON payments.subject_id = subjects.subject_id
        JOIN levels ON payments.level_id = levels.level_id
        WHERE payments.teacher_id = ?
        GROUP BY month_year
        ORDER BY month_year DESC
    """, (teacher_id,))
    monthly_payments = c.fetchall()
    db.close()

    return render_template('payments.html', payments=payments, monthly_payments=monthly_payments, teacher=teacher)


@app.route("/receipt/<student_id>/<subject_id>/<level_id>", methods=['GET', 'POST'])
def receipt(student_id, subject_id, level_id):

    # Get student data
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
    student = c.fetchone()
    db.close()

    if not student:
        flash("ไม่มีนักเรียนไอดีนี้", "warning")
        return redirect(url_for('students'))

    # Get registration data
    db = get_db()
    c = db.cursor()
    c.execute(
        """SELECT registrations.times, prices.price, students.f_name, students.student_id, students.title, students.l_name, students.n_name, subjects.subject_name, levels.level_name FROM registrations 
        JOIN students ON registrations.student_id = students.student_id
        JOIN subjects ON registrations.subject_id = subjects.subject_id
        JOIN levels ON registrations.level_id = levels.level_id
        JOIN prices ON subjects.subject_id = prices.subject_id and levels.level_id = prices.level_id
        WHERE registrations.student_id = ? and registrations.subject_id = ? and registrations.level_id = ?""", (student_id, subject_id, level_id))
    data = c.fetchone()
    db.close()

    if int(data[0]) % 4 == 0:
        payment = data[1] * (data[0] // 4)
        return render_template('receipt.html', data=data, payment=payment)
    else:
        flash("จำนวนครั้งหาร 4 ไม่ลงตัว", "warning")
        return redirect(url_for('students'))


@app.route('/export_excel')
def export_excel():
    # Connect to your SQLite database
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()

    # Create a dictionary to hold table queries
    tables = {
        'students': 'SELECT * FROM students',
        'teachers': 'SELECT * FROM teachers',
        'subjects': 'SELECT * FROM subjects',
        'levels': 'SELECT * FROM levels',
        'attendances': 'SELECT * FROM attendances',
        'registrations': 'SELECT * FROM registrations',
        'prices': 'SELECT * FROM prices',
        'payments': 'SELECT * FROM payments',
    }

    # Create a BytesIO buffer to hold the Excel file in memory
    output = BytesIO()

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, query in tables.items():
        # Execute the query and fetch data
        cursor.execute(query)
        rows = cursor.fetchall()

        # Get column names from the cursor
        columns = [description[0] for description in cursor.description]

        # Create a new sheet in the workbook
        sheet = workbook.create_sheet(title=sheet_name)

        # Write the column headers to the first row
        for col_num, column_title in enumerate(columns, 1):
            sheet.cell(row=1, column=col_num, value=column_title)

        # Write the data rows
        for row_num, row_data in enumerate(rows, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=cell_data)

    # Close the database connection
    conn.close()

    # Save the workbook to the BytesIO buffer
    workbook.save(output)

    # Prepare the response
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=output.xlsx"}
    )


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file and allowed_file(file.filename):
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

        # Import the file into the SQLite database
        import_excel_to_sqlite(file_path, 'database.db')

        flash('File successfully uploaded and imported')
        return redirect(url_for('students'))

    flash('Invalid file format')
    return redirect(request.url)


def import_excel_to_sqlite(excel_file_path, sqlite_db_path):
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)

    # Connect to the SQLite database
    conn = sqlite3.connect(sqlite_db_path)
    cursor = conn.cursor()

    # Loop through each sheet in the Excel file
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Get the headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Create a table with the same name as the sheet if it doesn't exist
        column_definitions = ', '.join(
            [f'"{header}" TEXT' for header in headers])
        create_table_query = f'CREATE TABLE IF NOT EXISTS "{sheet_name}" ({column_definitions})'
        cursor.execute(create_table_query)

        # Insert rows into the SQLite table
        for row in sheet.iter_rows(min_row=2, values_only=True):
            placeholders = ', '.join(['?' for _ in headers])

            # Using INSERT OR IGNORE to avoid UNIQUE constraint violations
            insert_query = f'INSERT OR IGNORE INTO "{sheet_name}" VALUES ({placeholders})'
            cursor.execute(insert_query, row)

    # Commit and close the connection
    conn.commit()
    conn.close()


@app.route('/generate-pdf', methods=['GET', 'POST'])
def generate_pdf():

    if request.method == 'POST':

        title = request.form['title']
        f_name = request.form['f_name']
        l_name = request.form['l_name']
        subject = request.form['subject']
        times = request.form['times']
        payment = request.form['payment']

        buffer = BytesIO()

        c = canvas.Canvas(buffer, pagesize=A4)
        c.setDash(1, 2)
        width, height = A4

        pdfmetrics.registerFont(TTFont('THSarabunNEW', os.path.join(
            os.path.dirname(__file__), 'fonts', 'THSarabunNew.ttf')))

        c.setFont("THSarabunNEW", 18)
        c.drawCentredString(width / 2, height - 100, "ใบเสร็จรับเงิน")

        c.setFont("THSarabunNEW", 10)
        c.drawCentredString(width / 2, height - 120, "RECEIPT")

        c.setFont("THSarabunNEW", 20)
        c.drawCentredString(width / 2, height - 140,
                            "โรงเรียนศุภนิจการดนตรีและภาษา Supanit Music & Language School")

        c.setFont("THSarabunNEW", 10)
        c.drawCentredString(width / 2, height - 160,
                            "888/5 หมู่ 4 ถนนวัชรพล แขวงคลองถนน เขตสายไหม กรุงเทพ 10220")
        c.drawCentredString(width / 2, height - 170, "โทร. 02-1530775-6")

        c.setFont("THSarabunNEW", 16)

        ################################################################################

        c.drawString(400, height - 220, "วันที่:")
        c.drawString(450, height - 220, date.today().strftime("%d/%m/%Y"))

        ################################################################################

        c.drawString(400, height - 240, "เลขที่:")
        c.line(440, height - 240 - 2, 540, height - 240 - 2)

        ################################################################################

        c.drawString(60, height - 260, "ได้รับเงินจาก:")
        c.drawString(120, height - 260, title + " " + f_name + " " + l_name)

        # Draw a dashed line
        c.line(120, height - 260 - 2, 260, height - 260 - 2)

        c.drawString(300, height - 260, "วิชา:")
        c.drawString(360, height - 260, subject)

        c.line(360, height - 260 - 2, 540, height - 260 - 2)

        ################################################################################

        c.drawString(60, height - 280, "ชำระสำหรับ:")
        c.drawString(120, height - 280, "ค่าเรียน" +
                     " " + times + " " + "ครั้ง")

        c.line(120, height - 280 - 2, 260, height - 280 - 2)

        c.drawString(300, height - 280, "จำนวนเงิน:")
        c.drawString(360, height - 280, payment + " " + "บาท")

        c.line(360, height - 280 - 2, 540, height - 280 - 2)

        c.setFont("THSarabunNEW", 16)
        c.drawCentredString(width / 2, height - 460,
                            "ใบเสร็จรับเงินที่ถูกต้อง จะต้องมีลายเซ็นของเจ้าหน้าที่ผู้รับมอบอำนาจ และประทับตราโรงเรียน")

        ################################################################################

        c.drawString(60, height - 320, "รับชำระ:")

        draw_checkbox(c, 120, height - 320, size=12, checked=False)
        c.drawString(140, height - 320, "บัตรเครดิต/Credit Card")

        draw_checkbox(c, 300, height - 320, size=12, checked=False)
        c.drawString(320, height - 320, "เงินสด/Cash")

        ################################################################################

        c.drawString(300, height - 400, "ผู้รับเงิน:")
        c.line(350, height - 400 - 2, 540, height - 400 - 2)

        ################################################################################

        c.setDash([])

        c.rect(50, height - 480, width - 80, 450, stroke=1, fill=0)

        c.showPage()
        c.save()

        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name='receipt.pdf', mimetype='application/pdf')


def draw_checkbox(c: canvas.Canvas, x, y, size=10, checked=False):

    c.setDash([])

    # Draw the box
    c.setLineWidth(1)
    c.rect(x, y, size, size)

    # Draw the checkmark if checked
    if checked:
        c.setLineWidth(1.5)
        c.line(x, y, x + size, y + size)
        c.line(x, y + size, x + size, y)

    c.setDash(1, 2)


def open_browser():
    webbrowser.open_new('http://127.0.0.1:5000/')


if __name__ == '__main__':
    threading.Timer(1, open_browser).start()
    app.run()
